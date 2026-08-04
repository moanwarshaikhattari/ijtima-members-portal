# myapp/views.py
import random
#excel export
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
#end excel export
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.contrib.auth.models import User
from django.conf import settings
from .models import Member
#whatsapp
import urllib.parse

def Userlogin(request):
    # 1. Agar user pehle se logged-in hai, toh unke role ke hisaab se redirect karein
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_staff:
            return redirect('allmembers')
        return redirect('members')

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        # 2. Email se Username find karein (Multiple users with same email handle karne ke liye filter.first())
        user_obj = User.objects.filter(email=email).first()

        if user_obj:
            user = authenticate(
                request, username=user_obj.username, password=password
            )

            if user is not None:
                login(request, user)

                # 3. Role-based Redirection (Admin vs Normal User)
                if user.is_superuser or user.is_staff:
                    return redirect('allmembers')  # Admin goes to allmembers
                else:
                    return redirect('members')  # Normal user goes to members

        # Log-in failed (Wrong email or password)
        messages.error(request, 'Invalid email or password.')

    data = {'title': 'Login - Members Portal'}
    return render(request, 'login.html', data)


def Usersignup(request):
    if request.user.is_authenticated:
        return redirect('members')

    # STEP 2: VERIFY OTP
    if request.method == 'POST' and 'verify_otp' in request.POST:
        entered_otp = request.POST.get('otp', '').strip()
        signup_data = request.session.get('signup_data')

        if not signup_data:
            messages.error(request, "Session expired. Please register again.")
            return redirect('register')

        if entered_otp == str(signup_data['otp']):
            # Create user in Django auth_user
            user = User.objects.create_user(
                username=signup_data['email'],
                email=signup_data['email'],
                password=signup_data['password'],
                first_name=signup_data['full_name']
            )
            
            # Save extra fields in User Profile
            user.profile.department = signup_data['department']
            user.profile.location = signup_data['location']
            user.profile.mobile = signup_data['mobile']
            user.profile.save()

            # Clean session data after successful registration
            del request.session['signup_data']
            
            messages.success(request, "Account created successfully! Please login.")
            return redirect('login')
        else:
            messages.error(request, "Invalid OTP code. Please try again.")
            return render(request, 'register.html', {
                'title': 'New Zimmedar Register',
                'otp_sent': True,
                'email': signup_data['email']
            })

    # STEP 1: SEND OTP
    elif request.method == 'POST' and 'send_otp' in request.POST:
        full_name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        department = request.POST.get('department', '').strip()
        location = request.POST.get('location', '').strip()
        password = request.POST.get('password', '').strip()

        # Check if user already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already registered.")
            return redirect('register')

        # Generate 6-digit random OTP
        otp = str(random.randint(100000, 999999))

        # Store form data & OTP in session
        request.session['signup_data'] = {
            'full_name': full_name,
            'email': email,
            'mobile': mobile,
            'department': department,
            'location': location,
            'password': password,
            'otp': otp
        }

        subject = 'Your Verification OTP Code'
        message = f'Hello {full_name},\n\nYour OTP code for registration is: {otp}'
        
        try:
            # Send email to the entered recipient
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            # messages.success(request, f"OTP code sent to {email}. Please enter it below.")
            
            return render(request, 'register.html', {
                'title': 'New Zimmedar Register',
                'otp_sent': True,
                'email': email
            })

        except Exception as e:
            messages.error(request, "Failed to send OTP. Please check your email configuration.")
            return redirect('register')

    return render(request, 'register.html', {
        'title': 'New Zimmedar Register', 
        'otp_sent': False
    })

@login_required(login_url='login')
def Userprofile(request):
    if not request.user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':
        request.user.first_name = request.POST.get('name')
        request.user.email = request.POST.get('email')
        request.user.save()

        # Update profile fields including mobile
        profile = request.user.profile
        profile.department = request.POST.get('department')
        profile.location = request.POST.get('location')
        profile.mobile = request.POST.get('mobile')  # <-- Save Mobile
        profile.save()

        messages.success(request, "Profile updated successfully!")
        return redirect('profile')

    return render(request, 'profile.html', {'title': 'Zimmedar Profile'})



def Userlogout(request):
    logout(request)
    return redirect('login')


def get_user_full_name(user):
    """Logged in user ka First & Last name fetch karne ke liye helper function"""
    full_name = user.get_full_name().strip()
    return full_name if full_name else user.username

@login_required(login_url='login')
def member_list(request):
    # Sirf logged-in user ke add kiye huye members dikhenge
    members = Member.objects.filter(created_by=request.user).order_by('-joined_at')
    # Agar Superuser/Admin hai toh SARE members milenge, nahi toh sirf Logged-In User ke
    # if request.user.is_superuser or request.user.is_staff:
    #     members = Member.objects.all().order_by('-joined_at')
    # else:
    #     members = Member.objects.filter(created_by=request.user).order_by('-joined_at')
    
    # Session se WhatsApp URL nikalein
    whatsapp_url = request.session.pop('open_whatsapp_url', None)

    return render(request, 'members.html', {
        'title': 'Members List',
        'members': members,
        'whatsapp_url': whatsapp_url
    })

@login_required(login_url='login')
def add_member(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        location = request.POST.get('location')
        blood = request.POST.get('blood')
        

        # CHECK: Entire Database me Mobile exist karta hai ya nahi? (Anwar ya kisi ne bhi add kiya ho)
        if Member.objects.filter(mobile=mobile).exists():
            # Error message pass karenge
            return render(request, 'add-member.html', {
                'title': 'Add Member',
                'error_mobile': 'This mobile number already exists in database!',
                'entered_name': name,
                'entered_mobile': mobile,
                'entered_location': location,
                'entered_blood':blood,
                "Member": Member,
            })

        # Save Member
        zimmedar = get_user_full_name(request.user)
        Member.objects.create(
            name=name,
            mobile=mobile,
            location=location,
            blood_donate=blood,
            zimmedar_name=zimmedar,
            created_by=request.user
        )
       # Build WhatsApp URL
        # Message Text
        message_body = f"Hello {name},\n\nWelcome! Your membership registration is successful.\nLocation: {location}"
        
        # FIX 1: safe='' set karna ZAROORI hai taaki \n breaks %0A ban jaye
        encoded_message = urllib.parse.quote(message_body, safe='')

        # FIX 2: Correct URL with exact parameters
        whatsapp_url = f"https://api.whatsapp.com/send?phone=91{mobile}&text={encoded_message}"

        request.session['open_whatsapp_url'] = whatsapp_url

        return redirect('members')

    return render(request, 'add-member.html', {'title': 'Add Member','Member': Member,})

@login_required(login_url='login')
def edit_member(request, pk):
    member = get_object_or_404(Member, pk=pk)

    if request.method == 'POST':
        member.name = request.POST.get('name')
        member.mobile = request.POST.get('mobile')
        member.location = request.POST.get('location')
        member.blood_donate = request.POST.get('blood')
        member.save()
        return redirect('members')

    return render(request, 'member-details.html', {
        'title': 'Edit Member Details',
        'member': member,
        "Member": Member,  # For LOCATION_CHOICES
    })


#export in excel file for member
@login_required(login_url='login')
def export_members_excel(request):
    # 1. Logged-in user ke specific members fetch karein
    if request.user.is_staff:
        members = Member.objects.all().order_by('-joined_at')
    else:
        members = Member.objects.filter(created_by=request.user).order_by('-joined_at')

    # 2. Workbook and Sheet Initialize Karein
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Members"

    # 3. Excel Sheet Header Styling (Added "Sr. No.")
    headers = [
        "Sr. No.",
        "Name",
        "Mobile Number",
        "Location",
        "Blood Donate",
        "Joined Date",
        "Zimmedar Name",
    ]
    ws.append(headers)

    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Filtered Data Rows Fill Karein (With Sr. No. counter)
    for idx, member in enumerate(members, start=1):
        joined_date = (
            member.joined_at.strftime("%b %d, %Y") if member.joined_at else ""
        )
        ws.append([
            idx,  # Serial Number added here
            member.name,
            member.mobile,
            member.location,
            member.blood_donate,
            joined_date,
            member.zimmedar_name,
        ])

    # 5. Column Width Adjustments
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 6. HttpResponse Return Karein
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{request.user.username}_members.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

#download excel file for registered users/zimmedar
@login_required(login_url='login')
def export_users_excel(request):
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registered Users'

    # Headers Styling
    header_fill = PatternFill(
        start_color='1F2937', end_color='1F2937', fill_type='solid'
    )
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')

    # Define Header Titles
    headers = [
        'SR No.',
        'Username',
        'Full Name',
        'Mobile',
        'Email',
        'Location',
        'Department',
        'Date Joined',
    ]
    ws.append(headers)

    # Apply style to headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Fetch User & Profile Data
    users = (
        User.objects.all().select_related('profile').order_by('-date_joined')
    )

    # Write Data Rows
    for idx, user_obj in enumerate(users, start=1):
        # Fallback handling for profile values
        profile = getattr(user_obj, 'profile', None)
        mobile = profile.mobile if profile and profile.mobile else '-'
        location = profile.location if profile and profile.location else '-'
        department = (
            profile.department if profile and profile.department else '-'
        )

        full_name = user_obj.get_full_name() or user_obj.username
        date_joined = (
            user_obj.date_joined.strftime('%d-%b-%Y')
            if user_obj.date_joined
            else '-'
        )

        row = [
            idx,
            user_obj.username,
            full_name.title(),
            mobile,
            user_obj.email or '-',
            location,
            department,
            date_joined,
        ]
        ws.append(row)

    # Auto-adjust column widths for neat look
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Prepare HTTP Response with Excel MIME type
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="Registered_Users_List.xlsx"'
    )

    wb.save(response)
    return response
#admin view pages
@login_required(login_url='login')
def all_members(request):
    # Sirf logged-in user ke add kiye huye members dikhenge
    # members = Member.objects.filter(created_by=request.user).order_by('-joined_at')
    members = Member.objects.all().order_by('-joined_at')
    # Agar Superuser/Admin hai toh SARE members milenge, nahi toh sirf Logged-In User ke
    # if request.user.is_superuser or request.user.is_staff:
    #     members = Member.objects.all().order_by('-joined_at')
    # else:
    #     members = Member.objects.filter(created_by=request.user).order_by('-joined_at')

    return render(request, 'allmembers.html', {
        'title': 'All Members List',
        'members': members,
    })

@login_required(login_url='login')
def all_zimmedar(request):
    # Saare registered Users aur unki Profile ka data ek saath fetch kar rahe hain
    users = (
        User.objects.all().select_related('profile').order_by('-date_joined')
    )

    data = {
        'title': 'Registered Users & Profiles',
        'users_list': users,  # Registered Users ki list
    }
    return render(request, 'allzimmedar.html', data)